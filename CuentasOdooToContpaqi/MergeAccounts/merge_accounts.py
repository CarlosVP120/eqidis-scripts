#!/usr/bin/env python3
"""
merge_accounts.py

Script para combinar dos archivos de cuentas contables.
Toma el primer archivo como base y agrega las cuentas del segundo archivo
que no existan en el primero, basándose en el código de cuenta.

Uso:
   source .venv/bin/activate && python3 MergeAccounts/merge_accounts.py MergeAccounts/entry_a.xlsx MergeAccounts/entry_b.xlsx MergeAccounts/output.xls

Uso con base CONTPAQI
   source .venv/bin/activate && python3 MergeAccounts/merge_accounts.py MergeAccounts/contpaqi_base.xlsx MergeAccounts/output.xlsx MergeAccounts/cuentas.xls

Requisitos:
  pip install pandas openpyxl xlwt
"""

import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import xlwt

def normalize_code(code):
    """
    Normaliza un código de cuenta para comparación.
    Remueve espacios y convierte a string.
    """
    if pd.isna(code) or code == "":
        return ""
    return str(code).strip()

def read_accounts_file(file_path):
    """
    Lee un archivo de cuentas y extrae las filas de datos (tipo 'C').
    Retorna un DataFrame con las cuentas y sus códigos normalizados.
    """
    print(f"📖 Leyendo archivo: {file_path}")
    
    # Leer el archivo completo
    df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
    
    # Filtrar solo las filas que son cuentas (tipo 'C')
    # Asumimos que la primera columna indica el tipo y la segunda el código
    account_rows = []
    codes_found = set()
    
    for i in range(len(df)):
        # Obtener el tipo (primera columna) y código (segunda columna)
        tipo = df.iat[i, 0] if len(df.columns) > 0 else ""
        codigo = df.iat[i, 1] if len(df.columns) > 1 else ""
        
        # Solo procesar filas de tipo 'C' (cuentas)
        if str(tipo).strip().upper() == 'C':
            codigo_normalizado = normalize_code(codigo)
            if codigo_normalizado and codigo_normalizado not in codes_found:
                codes_found.add(codigo_normalizado)
                # Guardar toda la fila
                row_data = []
                for j in range(len(df.columns)):
                    cell_value = df.iat[i, j] if j < len(df.columns) else ""
                    row_data.append(cell_value)
                
                account_rows.append({
                    'codigo': codigo_normalizado,
                    'row_data': row_data,
                    'original_index': i
                })
    
    print(f"   ✅ Encontradas {len(account_rows)} cuentas únicas")
    return account_rows, df

def merge_accounts(base_file, additional_file, output_file):
    """
    Combina dos archivos de cuentas.
    """
    print("🔄 Iniciando proceso de combinación...")
    
    # Leer archivo base
    base_accounts, base_df = read_accounts_file(base_file)
    base_codes = {acc['codigo'] for acc in base_accounts}
    
    # Leer archivo adicional
    additional_accounts, additional_df = read_accounts_file(additional_file)
    
    # Encontrar cuentas del archivo adicional que no están en el base
    new_accounts = []
    for acc in additional_accounts:
        if acc['codigo'] not in base_codes:
            new_accounts.append(acc)
    
    print(f"📊 Resumen:")
    print(f"   • Cuentas en archivo base: {len(base_accounts)}")
    print(f"   • Cuentas en archivo adicional: {len(additional_accounts)}")
    print(f"   • Cuentas nuevas a agregar: {len(new_accounts)}")
    
    # Crear el archivo resultado
    print(f"📝 Creando archivo resultado: {output_file}")
    
    # Leer plantilla base con openpyxl
    wb_base = load_workbook(base_file)
    ws_base = wb_base.active
    
    # Copiar todos los datos de la plantilla base
    template_data = []
    for row in ws_base.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value
            # Convertir fechas a formato compatible
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            row_data.append(value)
        if any(v is not None for v in row_data):
            template_data.append(row_data)
    
    # Cerrar el workbook de openpyxl
    wb_base.close()
    
    # Crear nuevo workbook en formato .xls
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sheet1')
    
    # Estilos para colores (xlwt tiene limitaciones, pero intentamos mantener la funcionalidad)
    green_style = xlwt.easyxf('pattern: pattern solid, fore_colour light_green')
    default_style = xlwt.easyxf()
    
    # Escribir datos de la plantilla base
    for row_idx, row_data in enumerate(template_data):
        for col_idx, value in enumerate(row_data):
            if value is not None:
                # Determinar el tipo de dato para xlwt
                if isinstance(value, (int, float)):
                    ws.write(row_idx, col_idx, value)
                else:
                    ws.write(row_idx, col_idx, str(value))
    
    # Agregar las cuentas nuevas después de la plantilla
    start_row = len(template_data)
    for i, acc in enumerate(new_accounts):
        row_data = acc['row_data']
        row_num = start_row + i
        
        for col_num, value in enumerate(row_data):
            if value is not None:
                # Usar estilo verde para las filas nuevas
                if isinstance(value, (int, float)):
                    ws.write(row_num, col_num, value, green_style)
                else:
                    ws.write(row_num, col_num, str(value), green_style)
    
    # Guardar en formato .xls
    wb.save(output_file)
    
    print(f"✅ Proceso completado exitosamente!")
    print(f"   📄 Archivo resultado guardado en: {output_file}")
    print(f"   🟢 Las {len(new_accounts)} cuentas nuevas están marcadas en verde")

def main():
    if len(sys.argv) != 4:
        print("❌ Error: Número incorrecto de argumentos")
        print()
        print("Uso:")
        print("  python3 merge_accounts.py archivo_base.xlsx archivo_adicional.xlsx archivo_resultado.xls")
        print()
        print("Descripción:")
        print("  • archivo_base.xlsx: Archivo que se usará como base")
        print("  • archivo_adicional.xlsx: Archivo del cual se tomarán las cuentas faltantes")
        print("  • archivo_resultado.xls: Archivo donde se guardará el resultado (formato Excel 97-2004)")
        print()
        print("El script combinará ambos archivos, manteniendo todas las cuentas del archivo base")
        print("y agregando solo las cuentas del archivo adicional que no existan en el base.")
        sys.exit(1)
    
    base_file = sys.argv[1]
    additional_file = sys.argv[2]
    output_file = sys.argv[3]
    
    try:
        merge_accounts(base_file, additional_file, output_file)
    except FileNotFoundError as e:
        print(f"❌ Error: No se pudo encontrar el archivo: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
