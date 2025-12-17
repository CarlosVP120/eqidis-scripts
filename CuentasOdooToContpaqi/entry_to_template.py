#!/usr/bin/env python3
"""
entry_to_template.py

Uso:
  ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate && python -m pip install -U pip pandas openpyxl && python3 entry_to_template.py template.xlsx entry.xlsx output.xlsx

Si ya tienes el entorno:
  source .venv/bin/activate && python3 entry_to_template.py template.xlsx entry.xlsx output.xlsx

Requisitos:
  pip install pandas openpyxl
(La plantilla debe estar en .xlsx para conservar encabezados/estilos.)
"""

import sys, re, os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

def find_columns(df):
    low = [str(c).lower().strip() for c in df.columns]
    code_idx = next((i for i,s in enumerate(low) if "cod" in s or "clave" in s), None)
    name_idx = next((i for i,s in enumerate(low) if "nombre" in s or "cuenta" in s), None)
    tipo_idx = next((i for i,s in enumerate(low) if "tipo" in s), None)
    if code_idx is None: code_idx = 0
    if name_idx is None: name_idx = 1 if df.shape[1] > 1 else 0
    return code_idx, name_idx, tipo_idx

def extract_code_from_name(name):
    if not isinstance(name, str): return None, name
    m = re.match(r'^\s*([0-9]+(?:\.[0-9]+)*)\s+(.+)$', name)
    if m: return m.group(1).strip(), m.group(2).strip()
    return None, name.strip()

def sanitize_code_str(s):
    if not s: return ""
    s = str(s).strip()
    s = re.sub(r'[^0-9\.]', '', s)
    s = re.sub(r'(?:\.0)+$', '', s)
    return s

# Variable configurable para el número total de dígitos
TOTAL_DIGITS = 8

def normalize_code(raw_code, total_digits=TOTAL_DIGITS):
    """
    Normaliza un código removiendo puntos y rellenando con ceros al final
    hasta completar el número total de dígitos especificado.
    
    Args:
        raw_code: Código original (ej: "1.2.3" o "123")
        total_digits: Número total de dígitos deseado (default: 11)
    
    Returns:
        Código normalizado con el número especificado de dígitos
    """
    if not raw_code:
        return "0" * total_digits
    
    # Remover puntos y caracteres no numéricos
    clean_code = re.sub(r'[^0-9]', '', str(raw_code))
    
    # Si está vacío después de limpiar, retornar ceros
    if not clean_code:
        return "0" * total_digits
    
    # Rellenar con ceros al final hasta completar total_digits
    return clean_code.ljust(total_digits, "0")


def build_catalog_rows(entry_path):
    df = pd.read_excel(entry_path, dtype=str, keep_default_na=False)
    code_idx, name_idx, tipo_idx = find_columns(df)
    
    # Leer archivo SAT.xlsx para obtener CtaMayor e IdAgrupadorSAT
    # El archivo SAT.xlsx debe estar en el mismo directorio que el script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sat_file_path = os.path.join(script_dir, 'SAT.xlsx')
    
    try:
        # Leer el archivo SAT directamente con openpyxl para preservar formato exacto
        wb_sat = load_workbook(sat_file_path, data_only=False)
        ws_sat = wb_sat.active
        
        # Encontrar las columnas en la primera fila
        header_row = 1
        nombre_col_idx = None
        nivel_col_idx = None
        codigo_col_idx = None
        
        for col in range(1, ws_sat.max_column + 1):
            cell_value = ws_sat.cell(row=header_row, column=col).value
            if cell_value:
                cell_str = str(cell_value).lower().strip()
                if 'nombre' in cell_str:
                    nombre_col_idx = col
                elif 'nivel' in cell_str:
                    nivel_col_idx = col
                elif 'codigo' in cell_str or 'código' in cell_str:
                    codigo_col_idx = col
        
        # Crear diccionario para búsqueda rápida: nombre -> (nivel, codigo)
        sat_lookup = {}
        for row in range(2, ws_sat.max_row + 1):  # Empezar desde la fila 2 (después del encabezado)
            # Leer Nombre
            nombre_sat = ""
            if nombre_col_idx:
                nombre_cell = ws_sat.cell(row=row, column=nombre_col_idx)
                if nombre_cell and nombre_cell.value is not None:
                    nombre_sat = str(nombre_cell.value).strip()
            
            # Leer Nivel
            nivel_sat = ""
            if nivel_col_idx:
                nivel_cell = ws_sat.cell(row=row, column=nivel_col_idx)
                if nivel_cell and nivel_cell.value is not None:
                    nivel_sat = str(nivel_cell.value).strip()
            
            # Leer Codigo preservando formato exacto
            codigo_sat = ""
            if codigo_col_idx:
                codigo_cell = ws_sat.cell(row=row, column=codigo_col_idx)
                if codigo_cell and codigo_cell.value is not None:
                    # Si es texto, leerlo directamente (preserva "10.10" como texto)
                    if codigo_cell.data_type == 's':
                        codigo_sat = str(codigo_cell.value).strip()
                    elif isinstance(codigo_cell.value, (int, float)):
                        # Si es numérico, usar el formato de la celda para preservar decimales
                        num_format = getattr(codigo_cell, 'number_format', '')
                        if num_format and '.' in num_format:
                            # Contar los decimales en el formato (ej: "0.00" tiene 2 decimales)
                            decimal_part = num_format.split('.')[-1].split(';')[0]
                            num_decimals = len([c for c in decimal_part if c in '0#?'])
                            if num_decimals > 0:
                                codigo_sat = f"{codigo_cell.value:.{num_decimals}f}"
                            else:
                                codigo_sat = str(int(codigo_cell.value))
                        else:
                            # Si no hay formato específico, convertir a string
                            codigo_sat = str(codigo_cell.value).strip()
                    else:
                        codigo_sat = str(codigo_cell.value).strip()
            
            if nombre_sat:
                sat_lookup[nombre_sat.strip().lower()] = (nivel_sat, codigo_sat)
        
        wb_sat.close()
    except Exception as e:
        print(f"⚠️ No se pudo leer el archivo SAT.xlsx: {e}")
        sat_lookup = {}
    
    # Obtener niveles de sangría desde la hoja original (columna de nombre)
    wb_src = load_workbook(entry_path, data_only=True)
    ws_src = wb_src.active
    indents = []
    for i in range(len(df)):
        # Asumimos encabezados en la primera fila -> datos comienzan en la 2
        row_num = i + 2
        col_num = (name_idx + 1) if name_idx is not None else 1
        cell = ws_src.cell(row=row_num, column=col_num)
        indent_val = cell.alignment.indent if cell and cell.alignment else 0
        raw_cell_text = str(cell.value) if cell and cell.value is not None else ""
        leading_spaces = len(raw_cell_text) - len(raw_cell_text.lstrip(' '))

        # Si Excel no provee indent (>0), estimar por espacios; base debe ser 1
        indent_val_excel = int(indent_val or 0)
        if indent_val_excel == 0 and leading_spaces > 0:
            indent_level = max(1, leading_spaces // 2)
        else:
            indent_level = indent_val_excel + 1  # shift para que base sea 1
        indents.append(indent_level)

    rows = []
    today = datetime.today().strftime("%Y%m%d")
    # Pilas por nivel de sangría para rastrear código y nombre del padre
    parent_code_by_indent = {}  # raw_code para CtaSup
    parent_name_by_indent = {}  # name para búsqueda SAT
    for i in range(len(df)):
        raw_code_cell = df.iat[i, code_idx] if code_idx < df.shape[1] else ""
        name_cell = df.iat[i, name_idx] if name_idx < df.shape[1] else ""
        tipo_cell = df.iat[i, tipo_idx] if (tipo_idx is not None and tipo_idx < df.shape[1]) else ""
        raw_code_cell = "" if pd.isna(raw_code_cell) else str(raw_code_cell).strip()
        name_cell = "" if pd.isna(name_cell) else str(name_cell).strip()
        tipo_cell = "" if pd.isna(tipo_cell) else str(tipo_cell).strip()

        raw_code = sanitize_code_str(raw_code_cell)
        name = name_cell

        if not raw_code:
            extracted, cleaned_name = extract_code_from_name(name_cell)
            if extracted:
                raw_code = sanitize_code_str(extracted)
                name = cleaned_name
            else:
                # no code anywhere -> ignorar
                continue
        else:
            # si el nombre repite el codigo al inicio, quitarlo
            m = re.match(r'^\s*' + re.escape(raw_code) + r'\s+(.+)$', name_cell)
            if m:
                name = m.group(1).strip()

        if not raw_code:
            continue

        indent_level = indents[i] if i < len(indents) else 1
        codigo_normalizado = normalize_code(raw_code)

        # calcular padre con base en la sangría (mínimo nivel = 1)
        parent_raw = parent_code_by_indent.get(indent_level - 1) if indent_level > 1 else None
        if parent_raw is None:
            cta_sup = "0" * TOTAL_DIGITS
        else:
            cta_sup = normalize_code(parent_raw)

        # actualizar pilas de padres para el siguiente nivel
        parent_code_by_indent[indent_level] = raw_code
        parent_name_by_indent[indent_level] = name
        # limpiar niveles más profundos que ya no aplican
        for k in list(parent_code_by_indent.keys()):
            if k > indent_level:
                del parent_code_by_indent[k]
        for k in list(parent_name_by_indent.keys()):
            if k > indent_level:
                del parent_name_by_indent[k]

        # Buscar en hoja SAT para obtener CtaMayor e IdAgrupadorSAT
        nombre_busqueda = name.strip().lower()
        if nombre_busqueda in sat_lookup:
            nivel_sat, codigo_sat = sat_lookup[nombre_busqueda]
            try:
                cta_mayor = int(nivel_sat) if nivel_sat else 2
            except (ValueError, TypeError):
                cta_mayor = 2
            id_agrupador = codigo_sat if codigo_sat else ""
            found_in_sat = True
        else:
            # No encontrado en SAT: usar valores por defecto
            cta_mayor = 2
            # Usar IdAgrupadorSAT del padre si existe
            parent_name = parent_name_by_indent.get(indent_level - 1) if indent_level > 1 else None
            if parent_name and parent_name.strip().lower() in sat_lookup:
                _, parent_codigo_sat = sat_lookup[parent_name.strip().lower()]
                id_agrupador = parent_codigo_sat if parent_codigo_sat else ""
                found_in_sat = False  # Yellow: not found but has parent
            else:
                id_agrupador = ""
                found_in_sat = None  # Red: not found and no parent

        rows.append({
            'data': [
                "C",
                codigo_normalizado,
                name,
                name,
                cta_sup,
                tipo_cell,
                0,
                cta_mayor,
                0,
                today,
                11,
                1,
                0,
                0,
                0,
                0,
                id_agrupador
            ],
            'found_in_sat': found_in_sat
        })
    return rows

def append_rows_to_template(template_path, rows, output_path):
    from openpyxl.styles import PatternFill
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Colores para diferentes casos
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # No encontrado pero tiene padre
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # No encontrado y sin padre
    
    for row_info in rows:
        row_data = row_info['data']
        found_in_sat = row_info['found_in_sat']
        
        # Agregar la fila
        ws.append(row_data)
        
        # Aplicar colores según el estado
        if found_in_sat is None:  # Red: not found and no parent
            row_num = ws.max_row
            for col in range(1, len(row_data) + 1):
                ws.cell(row=row_num, column=col).fill = red_fill
        elif found_in_sat is False:  # Yellow: not found but has parent
            row_num = ws.max_row
            for col in range(1, len(row_data) + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
        # found_in_sat is True: no color (found in SAT)
    
    wb.save(output_path)

def preprocess_entry_file(entry_path):
    """
    Preprocesa el archivo entry.xlsx:
    - Elimina las primeras 2 filas
    - Elimina las últimas 3 filas
    - Agrega columna I con fórmulas para determinar el tipo
    - Agrega columnas L y M con tabla de referencia
    """
    wb = load_workbook(entry_path, data_only=False)
    ws = wb.active
    
    # Validar que el archivo tenga suficientes filas
    if ws.max_row < 5:
        print(f"⚠️ Advertencia: El archivo tiene muy pocas filas ({ws.max_row}). No se puede preprocesar correctamente.")
        wb.close()
        return
    
    # Eliminar las últimas 3 filas primero
    for _ in range(3):
        if ws.max_row > 0:
            ws.delete_rows(ws.max_row)
    
    # Eliminar las primeras 2 filas
    if ws.max_row >= 2:
        ws.delete_rows(1, 2)
    
    # Agregar encabezado "Tipo" en columna I (columna 9)
    ws.cell(row=1, column=9, value="Tipo")
    
    # Agregar tabla de referencia en columnas L y M (columnas 12 y 13)
    referencia_data = [
        ("A", "Activo Deudora"),
        ("B", "Activo Acreedora"),
        ("C", "Pasivo Deudora"),
        ("D", "Pasivo Acreedora"),
        ("E", "Capital Deudora"),
        ("F", "Capital Acredora"),
        ("G", "Resultados Deudora"),
        ("H", "Resultados Acreedora"),
        ("K", "Orden Deudora"),
        ("L", "Orden Acreedora"),
    ]
    
    for i, (letra, descripcion) in enumerate(referencia_data, start=4):
        ws.cell(row=i, column=12, value=letra)  # Columna L
        ws.cell(row=i, column=13, value=descripcion)  # Columna M
    
    # Agregar fórmulas en columna I para cada fila de datos (empezando desde fila 2)
    for row_num in range(2, ws.max_row + 1):
        # Obtener el primer carácter del código (columna A) o nombre (columna B)
        cell_a = ws.cell(row=row_num, column=1)
        cell_b = ws.cell(row=row_num, column=2)
        
        codigo = str(cell_a.value) if cell_a.value else ""
        nombre = str(cell_b.value) if cell_b.value else ""
        
        # Determinar el primer dígito (de nombre o código)
        first_char = ""
        if nombre and nombre[0].isdigit():
            first_char = nombre[0]
        elif codigo and codigo[0].isdigit():
            first_char = codigo[0]
        
        # Construir la fórmula según el primer dígito
        if first_char in ["1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            digit = int(first_char)
            # Calcular las referencias L según el dígito
            # Para dígito 1: L4 (A), L5 (B)
            # Para dígito 2: L6 (C), L7 (D)
            # Para dígito 3: L8 (E), L9 (F)
            # Patrón: L(4 + 2*(digit-1)) para G>H, L(5 + 2*(digit-1)) para H>G
            l_deudora = 4 + 2 * (digit - 1)
            l_acreedora = 5 + 2 * (digit - 1)
            
            # Valor por defecto según el dígito
            default_value = "A" if digit == 1 else ("D" if digit == 2 else ("F" if digit == 3 else "A"))
            
            # Construir la fórmula siguiendo el patrón original
            # =IFNA(IF(MID(B2,1,1)="1",IFS(G2>H2,$L$4,H2>G2,$L$5),IF(MID(A2,1,1)="1",IFS(G2>H2,$L$4,H2>G2,$L$5))),"A")
            formula = f'=IFNA(IF(MID(B{row_num},1,1)="{first_char}",IFS(G{row_num}>H{row_num},$L${l_deudora},H{row_num}>G{row_num},$L${l_acreedora}),IF(MID(A{row_num},1,1)="{first_char}",IFS(G{row_num}>H{row_num},$L${l_deudora},H{row_num}>G{row_num},$L${l_acreedora}))),"{default_value}")'
        else:
            # Si no hay primer dígito, usar "A" por defecto
            formula = '="A"'
        
        # Asignar la fórmula a la celda I
        cell_i = ws.cell(row=row_num, column=9)
        cell_i.value = formula
    
    # Guardar el archivo procesado (sobrescribir)
    wb.save(entry_path)
    print(f"✅ Archivo entry.xlsx preprocesado: eliminadas 2 primeras y 3 últimas filas, agregadas columnas I, L y M")

def main(template_path, entry_path, output_path):
    # Preprocesar el archivo entry antes de procesarlo
    preprocess_entry_file(entry_path)
    
    rows = build_catalog_rows(entry_path)
    append_rows_to_template(template_path, rows, output_path)
    print(f"✅ Generado {output_path} con {len(rows)} filas tipo 'C'.")

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Uso: ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate && python -m pip install -U pip pandas openpyxl && python3 entry_to_template_v2.py template.xlsx entry.xlsx output.xlsx")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])