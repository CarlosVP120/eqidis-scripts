#!/usr/bin/env python3
"""
xml_to_contpaqi_xls.py

Uso:
  ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate && python -m pip install -U pip pandas openpyxl && python3 xml_to_contpaqi_xls.py template.xlsx entry.xml output.xlsx

Si ya tienes el entorno:
  source .venv/bin/activate && python3 xml_to_contpaqi_xls.py template.xlsx entry.xml output.xlsx

Explicación:
- Lee la plantilla (xls/xlsx) y genera un nuevo workbook .xlsx con las filas
  P / M1 / AM / AD con el mapeo que solicitaste.
"""
import sys
import os
import re
from pathlib import Path
import xml.etree.ElementTree as ET
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook

# --- helpers ---
def text_lower(x):
    return (x or "").lower()

# Variable configurable para el número total de dígitos
TOTAL_DIGITS = 8

def normalize_account_code(account_digits: str, total_digits=TOTAL_DIGITS) -> str:
    """
    Normaliza un código de cuenta removiendo puntos y rellenando con ceros al final
    hasta completar el número total de dígitos especificado.
    
    Args:
        account_digits: Código original de cuenta (ej: "1050103" o "1.05.01.03")
        total_digits: Número total de dígitos deseado (default: 11)
    
    Returns:
        Código normalizado con el número especificado de dígitos
    """
    if not account_digits:
        return "0" * total_digits
    
    # Remover puntos y caracteres no numéricos
    clean_code = re.sub(r'[^0-9]', '', str(account_digits))
    
    # Si está vacío después de limpiar, retornar ceros
    if not clean_code:
        return "0" * total_digits
    
    # Rellenar con ceros al final hasta completar total_digits
    return clean_code.ljust(total_digits, "0")


def safe_float(s):
    try:
        return float(s)
    except Exception:
        return 0.0

def truncate_referencia(referencia: str, max_length: int = 30) -> str:
    """
    Trunca la referencia a un máximo de caracteres para cumplir con las limitaciones de CONTPAQi.
    
    Args:
        referencia: Referencia original
        max_length: Longitud máxima permitida (default: 30)
    
    Returns:
        Referencia truncada si excede el límite
    """
    if not referencia:
        return ""
    
    if len(referencia) <= max_length:
        return referencia
    
    # Truncar a max_length caracteres
    return referencia[:max_length]

def determine_tipopol(num_un_iden: str, concepto_pol: str, transacciones: list):
    # transacciones: list of dicts with keys concept, descta, numcta, haber, debe
    n = (num_un_iden or "").upper()
    cpol = text_lower(concepto_pol)
    ttexts = [text_lower(t["Concepto"]) + " " + text_lower(t.get("DesCta","")) for t in transacciones]

    # si alguna transacción incluye "Cuenta transitoria" -> señal de que NO subir
    if any("cuenta transitoria" in t for t in ttexts):
        return None  # indica salto

    # reglas (ordenadas según tu lista, con pequeñas heurísticas)
    #  - NumUnIdenPol contiene "INV/" y "FACTU/" -> Tipo 3 (diario)
    if ("INV/" in n) or ("FACTU/" in n):
        return 3

    if "operaciones varias" in cpol:
        return 3

    # Effectively Paid rules
    if "effectively paid" in text_lower(concepto_pol):
        # check if any transaccion has IVA trasladado -> ingreso (1)
        if any("iva traslad" in t for t in ttexts):
            return 1
        if any("iva acredit" in t for t in ttexts):
            return 2

    # BNK logic and banco logic:
    if "BNK" in n or n.startswith("BNK"):
        # if any transaccion contains "clientes" -> ingreso
        if any("clientes" in t for t in ttexts):
            return 1
        # if any transaccion contains "proveedores" or "gastos" or "comisiones" -> egreso
        if any(("proveedores" in t or "gastos" in t or "comisiones" in t) for t in ttexts):
            return 2
        # banco + FACTU -> Egreso (2)
        if any(("banco" in t and "factu" in t) for t in ttexts):
            return 2
        # banco + INV -> Ingreso (1)
        if any(("banco" in t and "inv" in t) for t in ttexts):
            return 1
        # banco + SAT Impuestos -> Egreso (2)
        if any(("banco" in t and "sat impuestos" in t) for t in ttexts):
            return 2

    # If has "Banco" and "IVA acreditable" -> Egreso (2)
    if any(("banco" in t and "iva acredit" in t) for t in ttexts):
        return 2

    # fallback: diario
    return 3

# --- parse XML ---
def parse_xml_polizas(xml_path: str):
    # se asume el XML usa namespaces, reparo simple para encontrar "Poliza" y "Transaccion"
    tree = ET.parse(xml_path)
    root = tree.getroot()
    # obtener namespace map
    ns = {}
    # detect namespace prefix if present
    for k, v in root.attrib.items():
        if k.startswith("xmlns"):
            # ignore
            pass
    # para simplificar, buscaremos por tag suffix 'Poliza' etc
    polizas = []
    for p in root.findall(".//"):
        if p.tag.lower().endswith("poliza") and p.attrib.get("Fecha"):
            # direct Poliza element
            polizas.append(p)
    # if not found via suffix, try direct children
    if not polizas:
        for child in root:
            if child.tag.lower().endswith("poliza"):
                polizas.append(child)

    parsed = []
    for p in polizas:
        fecha = p.attrib.get("Fecha")
        concepto = p.attrib.get("Concepto","")
        num_un_iden = p.attrib.get("NumUnIdenPol","")
        transacciones = []
        for tx in p.findall(".//"):
            if tx.tag.lower().endswith("transaccion"):
                # read attributes
                tx_at = tx.attrib
                # parse nested CompNal if any
                compnals = []
                for c in tx:
                    if c.tag.lower().endswith("compnal"):
                        compnals.append(c.attrib)
                transacciones.append({
                    "Concepto": tx_at.get("Concepto",""),
                    "DesCta": tx_at.get("DesCta",""),
                    "NumCta": tx_at.get("NumCta",""),
                    "Haber": tx_at.get("Haber","0.00"),
                    "Debe": tx_at.get("Debe","0.00"),
                    "CompNal": compnals
                })
        parsed.append({
            "Fecha": fecha,
            "Concepto": concepto,
            "NumUnIdenPol": num_un_iden,
            "Transacciones": transacciones
        })
    return parsed

# --- build rows ---
def build_rows_from_parsed(polizas_parsed):
    rows = []
    folio = 1
    for pol in polizas_parsed:
        trans = pol["Transacciones"]
        # revisar si alguna transaccion contiene "Cuenta transitoria" -> skip poliza
        if any("cuenta transitoria" in text_lower(t["Concepto"]) or "cuenta transitoria" in text_lower(t.get("DesCta","")) for t in trans):
            # No se sube a CONTPAQi
            continue

        tipo = determine_tipopol(pol["NumUnIdenPol"], pol["Concepto"], trans)
        if tipo is None:
            # señal de skip
            continue

        # Fecha formato YYYYMMDD
        fecha = pol["Fecha"]
        try:
            # aceptar formatos YYYY-MM-DD o YYYY/MM/DD
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_fmt = dt.strftime("%Y%m%d")
        except Exception:
            # si viene sin guiones, intentar pasar como está
            fecha_fmt = (fecha or "").replace("-","").replace("/","")

        # construir P row
        # Orden: 'P', Fecha, TipoPol, Folio, Clase, IdDiario, Concepto, SistOrig, Impresa, Ajuste, Guid
        # Si el concepto es "Pago efectivo", solo usar NumUnIdenPol
        if pol['Concepto'].strip() == "Pago efectivo":
            concepto_poliza = pol['NumUnIdenPol']
        else:
            concepto_poliza = f"{pol['Concepto']} - {pol['NumUnIdenPol']}"
        
        p_row = [
            "P",
            fecha_fmt,
            tipo,
            folio,
            1,      # Clase siempre 1
            0,      # IdDiario siempre 0
            concepto_poliza,
            11,     # SistOrig
            0,      # Impresa
            0,      # Ajuste
            ""      # Guid vacío
        ]
        rows.append(p_row)

        # para AD final, recolectar UUIDs encontrados en la poliza
        ad_uuids = []

        # M1 rows
        for t in trans:
            numcta = t.get("NumCta","")
            # Normalizar cuenta usando TOTAL_DIGITS
            idcuenta = normalize_account_code(numcta)
            referencia = truncate_referencia(pol.get("NumUnIdenPol",""))
            haber = safe_float(t.get("Haber","0") or "0")
            debe = safe_float(t.get("Debe","0") or "0")
            # TipoMovto: Haber (negativo) -> ABONO (1). Debe (positivo) -> CARGO (0)
            if haber != 0 and abs(haber) > 0:
                tipomov = 1
                importe = abs(haber)
            else:
                tipomov = 0
                importe = abs(debe)

            # Limpiar concepto del movimiento eliminando "Pago efectivo - " si existe
            concepto_movimiento = t.get("Concepto","")
            if "Pago efectivo - " in concepto_movimiento:
                concepto_movimiento = concepto_movimiento.replace("Pago efectivo - ", "")
            
            m1_row = [
                "M1",
                idcuenta,
                referencia,
                tipomov,
                round(importe, 2),
                0,      # IdDiario
                0,      # ImporteME
                concepto_movimiento,
                "",     # IdSegNeg vacío
                "",     # Guid vacío
                ""      # FechaAplicacion vacío
            ]
            rows.append(m1_row)

            # si tiene CompNal con UUID_CFDI -> AM row for each CompNal
            comps = t.get("CompNal", [])
            for c in comps:
                uuid = c.get("UUID_CFDI") or c.get("UUID_CFDI".upper()) or c.get("UUID","")
                if not uuid:
                    # try other attr names:
                    uuid = c.get("UUID") or c.get("uuid") or c.get("UUID_CFDI".lower(), "")
                if uuid:
                    am_row = ["AM", uuid] + [""]*9  # pad to match columns (simple)
                    rows.append(am_row)
                    ad_uuids.append(uuid)

        # Al final de la poliza, agregar ADs (uno por uuid) -- como indicas, ADs van hasta el final
        for u in ad_uuids:
            ad_row = ["AD", u] + [""]*9
            rows.append(ad_row)

        folio += 1

    return rows

# --- escribir Excel resultado ---
def write_output_excel(template_path, rows, output_path):
    # Abrir plantilla (ya convertida a .xlsx)
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Buscar la última fila con datos (no vacía)
    last_row = 0
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            last_row = row[0].row
    
    # Insertar los datos inmediatamente después de la última fila con datos
    start_row = last_row + 1
    for i, row_data in enumerate(rows):
        row_num = start_row + i
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    wb.save(output_path)
    print(f"Wrote {output_path}")

# --- main ---
def main(template_path, xml_path, output_path):
    # parse xml
    parsed = parse_xml_polizas(xml_path)
    rows = build_rows_from_parsed(parsed)
    # Usar la plantilla como base y agregar los datos debajo de los encabezados
    write_output_excel(template_path, rows, output_path)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate && python -m pip install -U pip pandas openpyxl && python3 xml_to_contpaqi_xls.py template.xlsx entry.XML output.xlsx")
        sys.exit(1)
    plantilla = sys.argv[1]
    xmlf = sys.argv[2]
    out = sys.argv[3]
    main(plantilla, xmlf, out)
