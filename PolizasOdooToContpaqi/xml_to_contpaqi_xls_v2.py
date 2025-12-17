#!/usr/bin/env python3
"""
xml_to_contpaqi_xls_v2.py

Versión V2:
- Lee el XML de pólizas (como el generado por Odoo con mejoras de partner/ref).
- Lee un catálogo de cuentas en CSV (como "Grupos de cuentas.csv").
- Clasifica automáticamente cada póliza como:
    1 = Ingreso
    2 = Egreso
    3 = Diario
  usando SOLO la naturaleza contable de las cuentas (banco, clientes, proveedores, ingresos, gastos, etc.).

Uso:
  ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate \
    && python -m pip install -U pip pandas openpyxl xlwt \
    && python3 xml_to_contpaqi_xls_v2.py template.xlsx entry.xml "Grupos de cuentas.csv" output.xls

Si ya tienes el entorno:
  source .venv/bin/activate && python3 xml_to_contpaqi_xls_v2.py template.xlsx entry.xml "Grupos de cuentas.csv" output.xls
"""

import sys
import re
import xml.etree.ElementTree as ET
from datetime import datetime
import csv
from openpyxl import load_workbook
import xlwt


def text_lower(x):
    return (x or "").lower()


TOTAL_DIGITS = 8


def normalize_account_code(account_digits: str, total_digits=TOTAL_DIGITS) -> str:
    if not account_digits:
        return "0" * total_digits
    clean_code = re.sub(r"[^0-9]", "", str(account_digits))
    if not clean_code:
        return "0" * total_digits
    return clean_code.ljust(total_digits, "0")


def safe_float(s):
    try:
        return float(s)
    except Exception:
        return 0.0


def truncate_referencia(referencia: str, max_length: int = 30) -> str:
    if not referencia:
        return ""
    if len(referencia) <= max_length:
        return referencia
    return referencia[:max_length]


# =========================
# V2: Catálogo CSV de cuentas
# =========================

def load_account_groups_csv(csv_path: str):
    """
    Lee un catálogo de cuentas en CSV con columnas:
      - Fin de prefijo de código
      - Inicio de prefijo de código
      - Nombre

    Devuelve una lista de dicts con:
      - prefix_digits: prefijo numérico sin puntos (str)
      - name: nombre del grupo (str)
    """
    groups = []
    if not csv_path:
        return groups

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            start_prefix = row.get("Inicio de prefijo de código") or ""
            name = row.get("Nombre") or ""
            prefix_digits = re.sub(r"[^0-9]", "", start_prefix)
            if not prefix_digits:
                continue
            groups.append({
                "prefix_digits": prefix_digits,
                "name": name,
            })
    return groups


def infer_account_roles_from_group(prefix_digits: str, group_name: str):
    """
    A partir del prefijo numérico y el nombre del grupo del catálogo,
    devuelve un conjunto de "roles" para esa cuenta:
      - bank       (bancos / caja)
      - customer   (clientes / deudores comerciales)
      - supplier   (proveedores / acreedores comerciales)
      - income     (ingresos)
      - expense    (gastos / costos / resultado)
    """
    roles = set()
    name_lower = text_lower(group_name)
    first_digit = prefix_digits[0] if prefix_digits else ""

    if "banco" in name_lower or "bancos" in name_lower or "caja" in name_lower:
        roles.add("bank")
    if "clientes" in name_lower:
        roles.add("customer")
    if "proveedores" in name_lower or "acreedores" in name_lower:
        roles.add("supplier")
    if first_digit == "4" or "ingresos" in name_lower:
        roles.add("income")
    if first_digit in {"5", "6", "7"}:
        roles.add("expense")
    if "gastos" in name_lower or "costo" in name_lower:
        roles.add("expense")

    return roles


def build_account_role_index(groups):
    """
    Construye un índice rápido:
      - Para cada prefijo del catálogo, precalcula sus roles.
      - Luego se usará para buscar, dado un NumCta, cuál es el mejor match.
    """
    index = []
    for g in groups:
        prefix = g["prefix_digits"]
        name = g["name"]
        roles = infer_account_roles_from_group(prefix, name)
        index.append({
            "prefix_digits": prefix,
            "name": name,
            "roles": roles,
        })
    index.sort(key=lambda x: len(x["prefix_digits"]), reverse=True)
    return index


def get_roles_for_account(numcta: str, account_role_index):
    if not numcta:
        return set()
    clean_code = re.sub(r"[^0-9]", "", str(numcta))
    if not clean_code:
        return set()

    for item in account_role_index:
        pref = item["prefix_digits"]
        if clean_code.startswith(pref):
            return set(item["roles"])

    first_digit = clean_code[0]
    fallback_roles = set()
    if first_digit == "4":
        fallback_roles.add("income")
    elif first_digit in {"5", "6", "7"}:
        fallback_roles.add("expense")
    return fallback_roles


def determine_tipopol_v2(transacciones: list, account_role_index):
    """
    Determina el tipo de póliza usando SOLO la naturaleza contable
    de las cuentas (vía catálogo CSV).
    """
    has_bank = False
    has_customer = False
    has_supplier = False
    has_income = False
    has_expense = False

    for t in transacciones:
        numcta = t.get("NumCta", "")
        roles = get_roles_for_account(numcta, account_role_index)
        if "bank" in roles:
            has_bank = True
        if "customer" in roles:
            has_customer = True
        if "supplier" in roles:
            has_supplier = True
        if "income" in roles:
            has_income = True
        if "expense" in roles:
            has_expense = True

    if has_bank and (has_customer or has_income):
        return 1  # Ingreso
    if has_bank and (has_supplier or has_expense):
        return 2  # Egreso
    return 3  # Diario


# --- parse XML ---

def parse_xml_polizas(xml_path: str):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    polizas = []
    for p in root.findall(".//"):
        if p.tag.lower().endswith("poliza") and p.attrib.get("Fecha"):
            polizas.append(p)
    if not polizas:
        for child in root:
            if child.tag.lower().endswith("poliza"):
                polizas.append(child)

    parsed = []
    for p in polizas:
        fecha = p.attrib.get("Fecha")
        concepto = p.attrib.get("Concepto", "")
        num_un_iden = p.attrib.get("NumUnIdenPol", "")
        transacciones = []
        for tx in p.findall(".//"):
            if tx.tag.lower().endswith("transaccion"):
                tx_at = tx.attrib
                compnals = []
                for c in tx:
                    if c.tag.lower().endswith("compnal"):
                        compnals.append(c.attrib)
                transacciones.append({
                    "Concepto": tx_at.get("Concepto", ""),
                    "DesCta": tx_at.get("DesCta", ""),
                    "NumCta": tx_at.get("NumCta", ""),
                    "Haber": tx_at.get("Haber", "0.00"),
                    "Debe": tx_at.get("Debe", "0.00"),
                    "CompNal": compnals,
                })
        parsed.append({
            "Fecha": fecha,
            "Concepto": concepto,
            "NumUnIdenPol": num_un_iden,
            "Transacciones": transacciones,
        })
    return parsed


def build_rows_from_parsed_v2(polizas_parsed, account_role_index):
    rows = []
    folio = 1
    for pol in polizas_parsed:
        trans = pol["Transacciones"]
        if any("cuenta transitoria" in text_lower(t["Concepto"]) or
               "cuenta transitoria" in text_lower(t.get("DesCta", "")) for t in trans):
            continue

        tipo = determine_tipopol_v2(trans, account_role_index)

        fecha = pol["Fecha"]
        try:
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_fmt = dt.strftime("%Y%m%d")
        except Exception:
            fecha_fmt = (fecha or "").replace("-", "").replace("/", "")

        if pol.get("Concepto", "").strip() == "Pago efectivo":
            concepto_poliza = pol.get("NumUnIdenPol", "")
        else:
            concepto_poliza = f"{pol.get('Concepto','')} - {pol.get('NumUnIdenPol','')}"

        p_row = [
            "P",
            fecha_fmt,
            tipo,
            folio,
            1,
            0,
            concepto_poliza,
            11,
            0,
            0,
            "",
        ]
        rows.append(p_row)

        ad_uuids = []

        for t in trans:
            numcta = t.get("NumCta", "")
            idcuenta = normalize_account_code(numcta)
            referencia = truncate_referencia(pol.get("NumUnIdenPol", ""))
            haber = safe_float(t.get("Haber", "0") or "0")
            debe = safe_float(t.get("Debe", "0") or "0")
            if haber != 0 and abs(haber) > 0:
                tipomov = 1
                importe = abs(haber)
            else:
                tipomov = 0
                importe = abs(debe)

            concepto_movimiento = t.get("Concepto", "")
            if "Pago efectivo - " in concepto_movimiento:
                concepto_movimiento = concepto_movimiento.replace("Pago efectivo - ", "")

            m1_row = [
                "M1",
                idcuenta,
                referencia,
                tipomov,
                round(importe, 2),
                0,
                0,
                concepto_movimiento,
                "",
                "",
                "",
            ]
            rows.append(m1_row)

            comps = t.get("CompNal", [])
            for c in comps:
                uuid = c.get("UUID_CFDI") or c.get("UUID_CFDI".upper()) or c.get("UUID", "")
                if not uuid:
                    uuid = c.get("UUID") or c.get("uuid") or c.get("UUID_CFDI".lower(), "")
                if uuid:
                    am_row = ["AM", uuid] + [""] * 9
                    rows.append(am_row)
                    ad_uuids.append(uuid)

        for u in ad_uuids:
            ad_row = ["AD", u] + [""] * 9
            rows.append(ad_row)

        folio += 1

    return rows


def write_output_excel(template_path, rows, output_path):
    """
    Lee la plantilla (puede ser .xlsx) y escribe el resultado final en formato .xls (Excel 97-2004).
    """
    # Leer plantilla con openpyxl
    wb_template = load_workbook(template_path)
    ws_template = wb_template.active

    # Copiar datos de la plantilla
    template_data = []
    for row in ws_template.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value
            # Convertir fechas a formato compatible
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            row_data.append(value)
        if any(v is not None for v in row_data):
            template_data.append(row_data)

    # Crear nuevo workbook en formato .xls
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sheet1')

    # Escribir datos de la plantilla
    for row_idx, row_data in enumerate(template_data):
        for col_idx, value in enumerate(row_data):
            if value is not None:
                # Determinar el tipo de dato para xlwt
                if isinstance(value, (int, float)):
                    ws.write(row_idx, col_idx, value)
                else:
                    ws.write(row_idx, col_idx, str(value))

    # Agregar las nuevas filas después de la plantilla
    start_row = len(template_data)
    for i, row_data in enumerate(rows):
        row_num = start_row + i
        for col_num, value in enumerate(row_data):
            if value is not None:
                if isinstance(value, (int, float)):
                    ws.write(row_num, col_num, value)
                else:
                    ws.write(row_num, col_num, str(value))

    # Guardar en formato .xls
    wb.save(output_path)
    print(f"Wrote {output_path} (Excel 97-2004 format)")


def main_v2(template_path, xml_path, catalog_csv_path, output_path):
    groups = load_account_groups_csv(catalog_csv_path)
    account_role_index = build_account_role_index(groups)
    parsed = parse_xml_polizas(xml_path)
    rows = build_rows_from_parsed_v2(parsed, account_role_index)
    write_output_excel(template_path, rows, output_path)


if __name__ == "__main__":
    if len(sys.argv) != 5:
        print(
            "Uso V2 (con catálogo CSV):\n"
            "  ([ -d .venv ] || python3 -m venv .venv) && source .venv/bin/activate "
            "&& python -m pip install -U pip pandas openpyxl xlwt &&\n"
            "  python3 xml_to_contpaqi_xls_v2.py template.xlsx entry.xml \"Grupos de cuentas.csv\" output.xls"
        )
        sys.exit(1)

    plantilla = sys.argv[1]
    xmlf = sys.argv[2]
    catalog_csv = sys.argv[3]
    out = sys.argv[4]
    main_v2(plantilla, xmlf, catalog_csv, out)


